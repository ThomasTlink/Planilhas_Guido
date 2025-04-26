import pandas as pd
import unicodedata
import itertools
import random
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from colorama import init, Fore, Style
from datetime import datetime

def calcular_digito_verificador(ean12):
    soma = 0
    for i, digito in enumerate(ean12):
        n = int(digito)
        soma += n if i % 2 == 0 else n * 3
    resto = soma % 10
    return str((10 - resto) % 10)

def gerar_ean13(prefixos=['789', '790']):
    prefixo = random.choice(prefixos)
    corpo = ''.join(str(random.randint(0, 9)) for _ in range(9))
    ean12 = prefixo + corpo
    digito_verificador = calcular_digito_verificador(ean12)
    return ean12 + digito_verificador

init(autoreset=True)

def print_inicio():
    print(f"{Fore.CYAN}{'='*50}")
    print(f"{Fore.GREEN}INÍCIO DA EXECUÇÃO: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    print(f"{Fore.CYAN}{'='*50}{Style.RESET_ALL}")

def print_fim(path):
    print(f"{Fore.CYAN}{'='*50}")
    print(f"{Fore.GREEN}Planilha gerada com sucesso em: {Fore.YELLOW}{path}")
    print(f"{Fore.GREEN}Término: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    print(f"{Fore.CYAN}{'='*50}")

def print_colunas(df):
    print(f"{Fore.MAGENTA}Colunas encontradas no arquivo:")
    for col in df.columns:
        print(f"  - {Fore.YELLOW}{col}")

def aplicar_cores_intercaladas(path_arquivo):
    wb = load_workbook(path_arquivo)
    ws = wb.active

    fill_laranja = PatternFill(start_color="FFF4CC", end_color="FFF4CC", fill_type="solid")  # Laranja claro
    fill_branco = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Branco

    cor_atual = fill_branco

    for row in ws.iter_rows(min_row=2):  # Pula o cabeçalho
        codigo_pai = row[1].value  # Coluna 'Código'

        if '-' not in str(codigo_pai):  # Produto pai
            cor_atual = fill_laranja if cor_atual == fill_branco else fill_branco

        for cell in row:
            cell.fill = cor_atual

    wb.save(path_arquivo)

# Mapeamento categoria → peso, largura, altura, profundidade
dimensoes_categoria = {
    'acessórios' : (0.2, 8, 14, 20),
    'blazer': (1, 15, 25, 35),
    'bolsas': (1, 15, 25, 35),
    'bolsas femininas': (1, 15, 25, 35),
    'botas femininas': (1, 15, 25, 35),
    'botas masculinas': (1, 15, 25, 35),
    'calçados': (1, 15, 25, 35),
    'calçados femininos': (1, 15, 25, 35),
    'calçados masculinos': (1, 15, 25, 35),
    'calças e saias': (1, 15, 25, 35),
    'camisas': (1, 15, 25, 35),
    'carteiras e porta cartões': (0.2, 8, 14, 20),
    'casacos': (1, 15, 25, 35),
    'chapéus': (0.2, 8, 14, 20),
    'cintos': (0.1, 8, 14, 20),
    'corta ventos': (0.2, 15, 25, 35),
    'esportivo': (1, 15, 25, 35),
    'frasqueiras': (2, 32, 36, 37),
    'jaquetas': (1, 15, 25, 35),
    'luvas': (0.2, 8, 14, 20),
    'malas de viagem': (5, 32, 36, 37),
    'necessaries': (0.2, 9, 18, 28),
    'outras categorias': (1, 15, 25, 35),
    'pastas e mochilas': (1, 15, 37, 51),
    'sacolas de viagem': (2, 32, 36, 37),
    'tênis femininos': (1, 15, 25, 35),
    'tênis masculinos': (1, 15, 25, 35),
    'vestuário femininino': (1, 15, 25, 35),
}

def normalizar_colunas(df):
    df.columns = [unicodedata.normalize('NFC', col).strip() for col in df.columns]
    return df

def gerar_planilha_completa(entrada_path, saida_path):
    df = pd.read_excel(entrada_path)
    df = normalizar_colunas(df)
    df = df.fillna('')

    colunas_esperadas = {'Código', 'Descrição', 'NCM', 'Preço', 'Estoque', 'Preço de custo', 'Fornecedor', 'Cód. no fornecedor', 'Categoria', 'Tamanhos', 'Cores'}
    faltando = colunas_esperadas - set(df.columns)
    if faltando:
        raise ValueError(f"Colunas ausentes na planilha de entrada: {faltando}")
    if df.empty:
        raise ValueError(f"A planilha de entrada está vazia")
    
    linhas = []

    for _, row in df.iterrows():
        codigo_pai = str(row['Código']).strip()
        nome = row['Descrição']
        if pd.isna(codigo_pai) or pd.isna(nome):
            continue
        ncm = row['NCM']
        preco = row['Preço']
        preco_custo = row['Preço de custo']
        fornecedor = row['Fornecedor']
        cod_fornecedor = row['Cód. no fornecedor']
        categoria = row['Categoria'] 
        tamanhos = list(map(str.strip, str(row['Tamanhos']).split(',')))
        cores = list(map(str.strip, str(row['Cores']).split(',')))

        estoque_lista = list(map(str.strip, str(row['Estoque']).split(',')))
        index_estoque = 0

        peso, largura, altura, profundidade = dimensoes_categoria.get(str(categoria).strip().lower(), (1, 10, 10, 10))

        # Produto pai
        ean = gerar_ean13()
        linhas.append({
            'ID': '',
            'Código': codigo_pai,
            'Descrição': nome,
            'Unidade': 'UN',
            'NCM': ncm,
            'Origem': 0,
            'Preço': f"R$ {preco}",
            'Valor IPI fixo': '',
            'Observações': '',
            'Situação': 'Ativo',
            'Estoque': '',
            'Preço de custo': f"R$ {preco_custo}",
            'Cód. no fornecedor': cod_fornecedor,
            'Fornecedor': fornecedor,
            'Localização': '',
            'Estoque máximo': 0,
            'Estoque mínimo': 0,
            'Peso líquido (Kg)': peso,
            'Peso bruto (Kg)': peso,
            'GTIN/EAN': ean,
            'GTIN/EAN da Embalagem': ean,
            'Largura do produto': largura,
            'Altura do Produto': altura,
            'Profundidade do produto': profundidade,
            'Data Validade': '',
            'Descrição do Produto no Fornecedor': '',
            'Descrição Complementar': '',
            'Itens p/ caixa': 1,
            'Produto Variação': 'PRODUTO',
            'Tipo Produção': 'Terceiros',
            'Classe de enquadramento do IPI': '',
            'Código na Lista de Serviços': '',
            'Tipo do item': 'Mercadoria para Revenda',
            'Grupo de Tags/Tags': '',
            'Tributos': 0,
            'Código Pai': '',
            'Código Integração': '',
            'Grupo de produtos': '',
            'Marca': '',
            'CEST': '',
            'Volumes': 0,
            'Descrição Curta': '',
            'Cross-Docking': 0,
            'URL Imagens Externas': '',
            'Link Externo': '',
            'Meses Garantia no Fornecedor': '',
            'Clonar dados do pai': 'NÃO',
            'Condição do Produto': 'NOVO',
            'Frete Grátis': 'NÃO',
            'Número FCI': '',
            'Vídeo': '',
            'Departamento': '',
            'Unidade de Medida': 'Centímetro',
            'Preço de Compra': '',
            'Valor base ICMS ST para retenção': 0,
            'Valor ICMS ST para retenção': 0,
            'Valor ICMS próprio do substituto': 0,
            'Categoria do produto': categoria,
            'Informações Adicionais': '',
        })

        # Variações
        for cor, tamanho in itertools.product(cores, tamanhos):
            estoque_var = estoque_lista[index_estoque] if index_estoque < len(estoque_lista) else ''
            index_estoque += 1
            
            codigo_tam = str(tamanho).strip() if pd.notna(tamanho) and str(tamanho).strip().lower() != 'nan' else ""
            codigo_cor = str(cor).strip() if pd.notna(cor) and str(cor).strip().lower() != 'nan' else ""
            codigo_var_parts = [codigo_pai]
            if codigo_cor:
                codigo_var_parts.append(codigo_cor)
            if codigo_tam:
                codigo_var_parts.append(codigo_tam)

            codigo_var = "-".join(codigo_var_parts) 

            tamanho_str = f"Tamanho:{tamanho.strip()}" if pd.notna(tamanho) and str(tamanho).strip().lower() != 'nan' else ""
            cor_str = f"Cor:{str(cor).strip()}" if pd.notna(cor) and str(cor).strip().lower() != 'nan' else ""
            descricao_var = ";".join(part for part in [tamanho_str, cor_str] if part)

            ean = gerar_ean13()
            linhas.append({
                'ID': '',
                'Código': codigo_var,
                'Descrição': descricao_var,
                'Unidade': 'UN',
                'NCM': ncm,
                'Origem': 0,
                'Preço': f"R$ {preco}",
                'Valor IPI fixo': '',
                'Observações': '',
                'Situação': 'Ativo',
                'Estoque': estoque_var,
                'Preço de custo': f"R$ {preco_custo}",
                'Cód. no fornecedor': cod_fornecedor,
                'Fornecedor': fornecedor,
                'Localização': '',
                'Estoque máximo': 0,
                'Estoque mínimo': 0,
                'Peso líquido (Kg)': peso,
                'Peso bruto (Kg)': peso,
                'Link Externo': '',
                'GTIN/EAN': ean,
                'GTIN/EAN da Embalagem': ean,
                'Largura do produto': largura,
                'Altura do Produto': altura,
                'Profundidade do produto': profundidade,
                'Data Validade': '',
                'Descrição do Produto no Fornecedor': '',
                'Descrição Complementar': '',
                'Itens p/ caixa': 1,
                'Produto Variação': 'VARIAÇÃO',
                'Tipo Produção': 'Terceiros',
                'Classe de enquadramento do IPI': '',
                'Código na Lista de Serviços': '',
                'Tipo do item': 'Mercadoria para Revenda',
                'Grupo de Tags/Tags': '',
                'Tributos': 0,
                'Código Pai': codigo_pai,
                'Código Integração': 0,
                'Grupo de produtos': '',
                'Marca': '',
                'CEST': '',
                'Volumes': 0,
                'Descrição Curta': '',
                'Cross-Docking': 0,
                'URL Imagens Externas': '',
                'Link Externo': '',
                'Meses Garantia no Fornecedor': '',
                'Clonar dados do pai': 'SIM',
                'Condição do Produto': 'NOVO',
                'Frete Grátis': 'NÃO',
                'Número FCI': '',
                'Vídeo': '',
                'Departamento': '',
                'Unidade de Medida': 'Centímetro',
                'Preço de Compra': '',
                'Valor base ICMS ST para retenção': 0,
                'Valor ICMS ST para retenção': 0,
                'Valor ICMS próprio do substituto': 0,
                'Categoria do produto': categoria,
                'Informações Adicionais': '',
            })

    df_final = pd.DataFrame(linhas)
    df_final.to_excel(saida_path, index=False)
    print_inicio()
    print_colunas(df)
    print_fim(saida_path)


entrada = 'produtos_novos.xlsx'
saida = 'saída_formatada.xlsx'

saida = f'saída_formatada_{datetime.now().strftime("%d-%m-%Y_%H-%M")}.xlsx'

gerar_planilha_completa(entrada, saida)
aplicar_cores_intercaladas(saida)
