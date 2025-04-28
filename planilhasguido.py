import pandas as pd
import unicodedata
import itertools
import random
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from colorama import init, Fore, Style
from datetime import datetime
from pathlib import Path

# Inicializar colorama
init(autoreset=True)

# === Utilitários ===

def print_inicio():
    print(f"{Fore.CYAN}{'='*50}")
    print(f"{Fore.GREEN}INÍCIO DA EXECUÇÃO: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    print(f"{Fore.CYAN}{'='*50}{Style.RESET_ALL}")

def print_fim(path):
    print(f"{Fore.CYAN}{'='*50}")
    print(f"{Fore.GREEN}Planilha gerada com sucesso em: {Fore.YELLOW}{path}")
    print(f"{Fore.GREEN}Término: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    print(f"{Fore.CYAN}{'='*50}")

def print_colunas(df):
    print(f"{Fore.MAGENTA}Colunas encontradas no arquivo:")
    for col in df.columns:
        print(f"  - {Fore.YELLOW}{col}")

def normalizar_colunas(df):
    df.columns = [unicodedata.normalize('NFC', col).strip() for col in df.columns]
    return df

# === Cálculo de EAN ===

def calcular_digito_verificador(ean12):
    soma = sum(int(d) * (3 if i % 2 else 1) for i, d in enumerate(ean12))
    return str((10 - soma % 10) % 10)

def gerar_ean13(prefixos=['789', '790']):
    prefixo = random.choice(prefixos)
    corpo = ''.join(random.choices('0123456789', k=9))
    ean12 = prefixo + corpo
    return ean12 + calcular_digito_verificador(ean12)

# === Aplicar Cores ===

def aplicar_cores_intercaladas(path_arquivo):
    wb = load_workbook(path_arquivo)
    ws = wb.active
    fill_laranja = PatternFill(start_color="FFF4CC", end_color="FFF4CC", fill_type="solid")
    fill_branco = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    cor_atual = fill_branco

    for row in ws.iter_rows(min_row=2):
        codigo = row[1].value
        if '-' not in str(codigo):
            cor_atual = fill_laranja if cor_atual == fill_branco else fill_branco
        for cell in row:
            cell.fill = cor_atual

    wb.save(path_arquivo)

# === Dados de Categorias ===

dimensoes_categoria = {
    'acessórios': (0.2, 8, 14, 20),
    'blazer': (1, 15, 25, 35),
    'bolsas': (1, 15, 25, 35),
    'botas femininas': (1, 15, 25, 35),
    'calçados': (1, 15, 25, 35),
    'calças e saias': (1, 15, 25, 35),
    'camisas': (1, 15, 25, 35),
    'carteiras e porta cartões': (0.2, 8, 14, 20),
    'chapéus': (0.2, 8, 14, 20),
    'cintos': (0.1, 8, 14, 20),
    'corta ventos': (0.2, 15, 25, 35),
    'esportivo': (1, 15, 25, 35),
    'frasqueiras': (2, 32, 36, 37),
    'jaquetas': (1, 15, 25, 35),
    'luvas': (0.2, 8, 14, 20),
    'malas de viagem': (5, 32, 36, 37),
    'necessaries': (0.2, 9, 18, 28),
    'outras categorias': (1, 15, 25, 35),
    'pastas e mochilas': (1, 15, 37, 51),
    'sacolas de viagem': (2, 32, 36, 37),
    'tênis femininos': (1, 15, 25, 35),
    'tênis masculinos': (1, 15, 25, 35),
    'vestuário femininino': (1, 15, 25, 35),
}

# === Função para criar produto (pai ou variação) ===

def criar_produto(codigo, descricao, ncm, preco, preco_custo, fornecedor, cod_fornecedor,
                  categoria, estoque, ean, peso, largura, altura, profundidade,
                  variacao=False, codigo_pai='', cor='', tamanho=''):
    
    descricao_final = descricao
    if variacao:
        descricao_final = ";".join(f"{k}:{v.strip()}" for k, v in [('Tamanho', tamanho), ('Cor', cor)] if v.strip())

    return {
        'ID': '',
        'Código': codigo,
        'Descrição': descricao_final,
        'Unidade': 'UN',
        'NCM': ncm,
        'Origem': 0,
        'Preço': f"R$ {preco}",
        'Valor IPI fixo': '',
        'Observações': '',
        'Situação': 'Ativo',
        'Estoque': estoque,
        'Preço de custo': f"R$ {preco_custo}",
        'Cód. no fornecedor': cod_fornecedor,
        'Fornecedor': fornecedor,
        'Localização': '',
        'Estoque máximo': 0,
        'Estoque mínimo': 0,
        'Peso líquido (Kg)': peso,
        'Peso bruto (Kg)': peso,
        'GTIN/EAN': ean,
        'GTIN/EAN da Embalagem': ean,
        'Largura do produto': largura,
        'Altura do Produto': altura,
        'Profundidade do produto': profundidade,
        'Data Validade': '',
        'Descrição do Produto no Fornecedor': '',
        'Descrição Complementar': '',
        'Itens p/ caixa': 1,
        'Produto Variação': 'VARIAÇÃO' if variacao else 'PRODUTO',
        'Tipo Produção': 'Terceiros',
        'Classe de enquadramento do IPI': '',
        'Código na Lista de Serviços': '',
        'Tipo do item': 'Mercadoria para Revenda',
        'Grupo de Tags/Tags': '',
        'Tributos': 0,
        'Código Pai': codigo_pai if variacao else '',
        'Código Integração': 0,
        'Grupo de produtos': '',
        'Marca': '',
        'CEST': '',
        'Volumes': 0,
        'Descrição Curta': '',
        'Cross-Docking': 0,
        'URL Imagens Externas': '',
        'Link Externo': '',
        'Meses Garantia no Fornecedor': '',
        'Clonar dados do pai': 'SIM' if variacao else 'NÃO',
        'Condição do Produto': 'NOVO',
        'Frete Grátis': 'NÃO',
        'Número FCI': '',
        'Vídeo': '',
        'Departamento': '',
        'Unidade de Medida': 'Centímetro',
        'Preço de Compra': '',
        'Valor base ICMS ST para retenção': 0,
        'Valor ICMS ST para retenção': 0,
        'Valor ICMS próprio do substituto': 0,
        'Categoria do produto': categoria,
        'Informações Adicionais': '',
    }

# === Processamento principal ===

def gerar_planilha_completa(entrada_path, saida_path):
    df = pd.read_excel(entrada_path)
    df = normalizar_colunas(df).fillna('')

    colunas_esperadas = {'Código', 'Descrição', 'NCM', 'Preço', 'Estoque', 'Preço de custo', 'Fornecedor', 'Cód. no fornecedor', 'Categoria', 'Tamanhos', 'Cores'}
    faltando = colunas_esperadas - set(df.columns)
    if faltando:
        raise ValueError(f"Colunas ausentes: {', '.join(faltando)}")
    if df.empty:
        raise ValueError("A planilha de entrada está vazia.")

    linhas = []

    for _, row in df.iterrows():
        codigo_pai = str(row['Código']).strip()
        if not codigo_pai or pd.isna(row['Descrição']):
            continue

        preco, preco_custo = row['Preço'], row['Preço de custo']
        fornecedor, cod_forn = row['Fornecedor'], row['Cód. no fornecedor']
        categoria = str(row['Categoria']).strip().lower()
        tamanhos = list(map(str.strip, str(row['Tamanhos']).split(';')))
        cores = list(map(str.strip, str(row['Cores']).split(';')))
        estoque = row['Estoque']
        ean = gerar_ean13()

        # Adiciona o produto pai
        linhas.append(criar_produto(codigo_pai, row['Descrição'], row['NCM'], preco, preco_custo, fornecedor, cod_forn, categoria, estoque, ean, 1, 10, 10, 10))

        # Adiciona as variações
        for cor, tamanho in itertools.product(cores, tamanhos):
            codigo_variacao = f"{codigo_pai}-{cor}-{tamanho}"
            linhas.append(criar_produto(codigo_variacao, row['Descrição'], row['NCM'], preco, preco_custo, fornecedor, cod_forn, categoria, estoque, ean, 1, 10, 10, 10, True, codigo_pai, cor, tamanho))

    # Salva a planilha
    df_final = pd.DataFrame(linhas)
    df_final.to_excel(saida_path, index=False)
    aplicar_cores_intercaladas(saida_path)

# === Execução ===

entrada = Path('produtos_novos.xlsx')
saida = Path(f"saída_formatada_{datetime.now().strftime('%d-%m-%Y_%H-%M-%S')}.xlsx")

print_inicio()
gerar_planilha_completa(entrada, saida)
print_fim(saida)
